import openpyxl
from utilities.Read_Config_File import Read_Confi_File


class Read_Excel_File:

    @staticmethod
    def read_excel(rownum, colnum):
        R = Read_Confi_File()
        Excel = openpyxl.load_workbook(R.Excel_Path())
        Sheet = Excel.active
        return Sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def write_excel(rownum, colnum, data):
        R = Read_Confi_File()
        Excel = openpyxl.load_workbook(R.Excel_Path())
        Sheet = Excel.active
        Sheet.cell(row=rownum, column=colnum).value = data
        Excel.save(R.Excel_Path())